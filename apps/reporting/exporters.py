import csv
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any, Callable, Iterable

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


Accessor = str | Callable[[Any], Any]


@dataclass(frozen=True)
class ReportColumn:
    header: str
    accessor: Accessor
    width: int = 18

    def value_from(self, row: Any) -> Any:
        if callable(self.accessor):
            return self.accessor(row)
        value = row
        for part in self.accessor.split("."):
            value = getattr(value, part, "")
            if value is None:
                return ""
        return value


@dataclass(frozen=True)
class TabularReport:
    title: str
    filename: str
    columns: tuple[ReportColumn, ...]
    rows: Iterable[Any]
    sheet_name: str = "Reporte"

    def iter_rows(self):
        for row in self.rows:
            yield [normalize_cell(column.value_from(row)) for column in self.columns]


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "isoformat") and value.__class__.__name__ in {"date", "time"}:
        return value.isoformat()
    if hasattr(value, "strftime") and value.__class__.__name__ == "datetime":
        return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")
    return str(value) if not isinstance(value, (int, float, bool)) else value


class BaseReportExporter:
    content_type = "application/octet-stream"
    extension = "bin"

    def __init__(self, report: TabularReport):
        self.report = report

    def response(self) -> HttpResponse:
        response = HttpResponse(self.render(), content_type=self.content_type)
        response["Content-Disposition"] = (
            f'attachment; filename="{self.report.filename}.{self.extension}"'
        )
        return response

    def render(self):
        raise NotImplementedError


class CsvReportExporter(BaseReportExporter):
    content_type = "text/csv; charset=utf-8"
    extension = "csv"

    def render(self):
        response = HttpResponse(content_type=self.content_type)
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow([column.header for column in self.report.columns])
        writer.writerows(self.report.iter_rows())
        return response.content


class ExcelReportExporter(BaseReportExporter):
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def render(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.report.sheet_name[:31] or "Reporte"
        headers = [column.header for column in self.report.columns]
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in self.report.iter_rows():
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(self.report.columns, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = column.width
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


class PdfReportExporter(BaseReportExporter):
    content_type = "application/pdf"
    extension = "pdf"

    def render(self):
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=24,
            bottomMargin=24,
        )
        styles = getSampleStyleSheet()
        data = [[column.header for column in self.report.columns], *list(self.report.iter_rows())]
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9DEE3")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FB")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        document.build([Paragraph(self.report.title, styles["Title"]), Spacer(1, 12), table])
        return buffer.getvalue()


EXPORTERS = {
    "csv": CsvReportExporter,
    "xlsx": ExcelReportExporter,
    "pdf": PdfReportExporter,
}


def supported_export_formats() -> set[str]:
    return set(EXPORTERS)


def export_report(report: TabularReport, file_format: str) -> HttpResponse:
    exporter_class = EXPORTERS[file_format]
    return exporter_class(report).response()
